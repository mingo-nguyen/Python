import openpyxl
from bs4 import BeautifulSoup
import json
import os
import pandas as pd

outputfile = "temp\\person_details.json"
if os.path.exists(outputfile):
    os.remove(outputfile)
    print(f"File '{outputfile}' deleted successfully.")


with open('temp\\details.bin', 'rb') as file:
    binary_content = file.read()
if binary_content == "Error":
    with open(outputfile, "w") as file:
        file.write("Error")
        print("Error: No Details HTML Page")
    pass
# Parse the binary content with BeautifulSoup
soup = BeautifulSoup(binary_content, "html.parser")
# get full name
find_nodes = soup.find_all(class_="name-given")
strFullName = find_nodes[0].text.strip()
# get address
find_nodes = soup.find_all(class_="address-current")
strAddress = find_nodes[0].text.strip()
# get wireless
find_nodes = soup.find_all("a", class_="phone", title=lambda value: value and ("Wireless" in value or "VOIP" in value), href=lambda value: value and "phone" in value)
arrWirelesses =["None found","None found","None found", "None found"]
for index, item in enumerate(find_nodes):
    if(index==4):
        break
    arrWirelesses[index] = item.text.strip()
strWireless1 = arrWirelesses[0]
strWireless2 = arrWirelesses[1]
strWireless3 = arrWirelesses[2]
strWireless4 = arrWirelesses[3]

#get Emails
find_nodes = soup.find_all("a", class_="email", title=lambda value: value and "Find other people" in value)
arrEmails =["None found","None found","None found"]
for index, item in enumerate(find_nodes):
    if(index==3):
        break
    arrEmails[index] = item.text.strip()
strEmail1 = arrEmails[0]
strEmail2 = arrEmails[1]
strEmail3 = arrEmails[2]

# get landline
find_nodes = soup.find_all("a", class_="phone", title=lambda value: value and "LandLine" in value, href=lambda value: value and "phone" in value)
arrLandlines =["None found","None found","None found"]
for index, item in enumerate(find_nodes):
    if(index==3):
        break
    arrLandlines[index] = item.text.strip()

strLandline1 = arrLandlines[0]
strLandline2 = arrLandlines[1]
strLandline3 = arrLandlines[2]

data = {
    "strPropertyAddress": "",
    "strAddress": strAddress,
    "strFullName": strFullName,
    "strEmail1": strEmail1,
    "strEmail2": strEmail2,
    "strEmail3": strEmail3,
    "strWireless1": strWireless1,
    "strWireless2": strWireless2,
    "strWireless3": strWireless3,
    "strWireless4": strWireless4,
    "strLandline1": strLandline1,
    "strLandline2": strLandline2,
    "strLandline3": strLandline3

}

# Define the filename for the JSON file


# Write the data to the JSON file
with open(outputfile, "w") as file:
    json.dump(data, file, indent=4)
print("Info written to", outputfile)



# Open the workbook
filename = "Input.xlsx"
wb = openpyxl.load_workbook(filename)

# Select the worksheet
sheet = wb["Master"]

# Append data from the dictionary to the next available row
next_row = sheet.max_row + 1

# Write data to the cells
for col_num, value in enumerate(data.values(), start=1):
    sheet.cell(row=next_row, column=col_num).value = value

# Save the workbook
wb.save(filename)